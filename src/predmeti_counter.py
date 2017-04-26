from collections import defaultdict
from openpyxl import load_workbook
predmeti = defaultdict(int)
from openpyxl import Workbook
wb = load_workbook("../podatki/izbirci_correct.xlsx")
sr= wb["correcttxt"]
c=2
while c!=77747:
    predmet=sr['C'+str(c)].value
    predmeti[predmet]+=1
    c+=1
#TODO, to ni cist pravilno ker se veckrat ponovijo predmeti na solo
wb_new = Workbook()
dest="../nasi_podatki/statistika_predmetov.xlsx"
ws1= wb_new.active
ws1.title = "St. predmetov nasploh"
ws1['A1']="Predmet"
ws1['B1']="Pojavitev"
counter=2
for key,value in predmeti.items():
    ws1['A'+str(counter)]=key
    ws1['B'+str(counter)]=value
    print(key,value)
    counter+=1
wb_new.save(filename=dest)