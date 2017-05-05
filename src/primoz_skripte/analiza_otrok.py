from collections import defaultdict
from openpyxl import load_workbook
predmeti = defaultdict(int)
from openpyxl import Workbook
wb = load_workbook("../../nasi_podatki/primoz/oddelki_otroc.xlsx")
sr= wb["List1"]
leto_ucenci=defaultdict(int)
leto_ucenke=defaultdict(int)
leto_romi=defaultdict(int)
counter=2
#109541
#12 - '1. razred'
#13 - '2. razred'
#14 - '3. razred'
#15 - '4. razred'
#16 - '5. razred'
#17 - '6. razred'
#18 - '7. razred'
#19 - '8. razred'
#20 - '9. razred'

while counter!=85764:
    razred = sr["A"+str(counter)].value
    ucenci = sr["C"+str(counter)].value
    ucenke = sr["D"+str(counter)].value
    romi = sr["E"+str(counter)].value
    if razred==12:
        leto_ucenci["1. razred"]+=int(ucenci)
        leto_ucenke["1. razred"]+= int(ucenke)
        leto_romi["1. razred"]+= int(romi)
    elif razred==13:
        leto_ucenci["2. razred"] += int(ucenci)
        leto_ucenke["2. razred"] += int(ucenke)
        leto_romi["2. razred"] += int(romi)
    elif razred == 14:
        leto_ucenci["3. razred"] += int(ucenci)
        leto_ucenke["3. razred"] += int(ucenke)
        leto_romi["3. razred"] += int(romi)
    elif razred == 15:
        leto_ucenci["4. razred"] += int(ucenci)
        leto_ucenke["4. razred"] += int(ucenke)
        leto_romi["4. razred"] += int(romi)
    elif razred == 16:
        leto_ucenci["5. razred"] += int(ucenci)
        leto_ucenke["5. razred"] += int(ucenke)
        leto_romi["5. razred"] += int(romi)
    elif razred == 17:
        leto_ucenci["6. razred"] += int(ucenci)
        leto_ucenke["6. razred"] += int(ucenke)
        leto_romi["6. razred"] += int(romi)
    elif razred == 18:
        leto_ucenci["7. razred"] += int(ucenci)
        leto_ucenke["7. razred"] += int(ucenke)
        leto_romi["7. razred"] += int(romi)
    elif razred == 19:
        leto_ucenci["8. razred"] += int(ucenci)
        leto_ucenke["8. razred"] += int(ucenke)
        leto_romi["8. razred"] += int(romi)
    elif razred == 20:
        leto_ucenci["9. razred"] += int(ucenci)
        leto_ucenke["9. razred"] += int(ucenke)
        leto_romi["9. razred"] += int(romi)
    counter+=1

print(leto_ucenci)
print(leto_ucenke)
print(leto_romi)

wb_new = Workbook()

dest="../../nasi_podatki/primoz/oddelki_populacija.xlsx"
ws1= wb_new.active
ws1.title = "St. predmetov nasploh"
ws1['A1']="Oddelek"
ws1['B1']="Ucenci"
ws1['C1']="Ucenke"
ws1['D1']="Romi"
counter=2
for key,value in leto_ucenci.items():
    ws1['A'+str(counter)]=key
    ws1['B'+str(counter)]=value
    counter+=1

counter=2
for key,value in leto_ucenke.items():
    ws1['C'+str(counter)]=value
    counter+=1

counter=2
for key,value in leto_romi.items():
    ws1['D'+str(counter)]=value
    counter+=1
wb_new.save(filename=dest)