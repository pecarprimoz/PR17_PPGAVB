from collections import defaultdict
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
wb = load_workbook("../../nasi_podatki/primoz/nb_sole_predmet.xlsx")
sr= wb["nb"]

#
vsi_predmeti = set()
vse_sole=dict()
regresija_podatki=defaultdict(lambda : defaultdict(int))
katere_predmete_sola=defaultdict(list)

## SIFRA TO REGIJA PRESLIKAVA ##
sifra_v_regijo=dict()
for i in range(2,774):
    sira = int(sr["D"+str(i)].value)
    sifra_v_regijo[sira]=sr["E"+str(i)].value

#print(sifra_v_regijo) is correct

## PREDMET ##
for i in range(2,77478):
    sola = sr["A" + str(i)].value
    predmet = sr["B"+str(i)].value
    vsi_predmeti.add(predmet)
    if predmet not in katere_predmete_sola[sola]:
        katere_predmete_sola[sola].append(predmet)
        temp_dict=defaultdict(int)
        regresija_podatki[sola][predmet]=sr["F"+str(i)].value

print(regresija_podatki)
#print(katere_predmete_sola)
## SOLA ##
for i in range(2,77478):
    sola = sr["A"+str(i)].value
    if sola not in vse_sole:
        vse_sole[sr["C"+str(i)].value]=sola

#print(vsi_predmeti)
#print(vse_sole)



wb_new = Workbook()
ws1= wb_new.active
dest="../../nasi_podatki/primoz/nb_model_current_regression.xlsx"
ws1.title = "Matrika solaXpredmet"
zgornji_counter=4
for ind,predmet in enumerate(list(vsi_predmeti)):
    ws1[get_column_letter(zgornji_counter)+str(1)].value=predmet
    zgornji_counter+=1

zgornji_counter=2
for key,value in vse_sole.items():
    if key in sifra_v_regijo:
        ws1["A"+str(zgornji_counter)]=sifra_v_regijo[key]
        ws1["B" +str(zgornji_counter)] = key
    ws1["C" +str(zgornji_counter)] = value
    zgornji_counter+=1



wb_new.save(dest)


wb_my_own = load_workbook("../../nasi_podatki/primoz/nb_model_current_regression.xlsx")
sr= wb_my_own["Matrika solaXpredmet"]
wb_new_two = Workbook()
ws_new= wb_new_two.active
dest="../../nasi_podatki/primoz/regr_model_final.xlsx"
ws_new.title = "Numberos"

#2C DO C528
for i in range(4,120): #ZGORN STOLPEC PREDMETU
    for j in range(2,528): #VSE SOLE, GLEJ C
        cur_sola=sr["C"+str(j)].value
        predmeti_sole=katere_predmete_sola[cur_sola]
        predmet_v_stolpcu=sr[get_column_letter(i)+str(1)].value
        print(predmet_v_stolpcu)
        if predmet_v_stolpcu in predmeti_sole:
            ws_new[get_column_letter(i)+str(j)]=regresija_podatki[cur_sola][predmet_v_stolpcu]
        else:
            ws_new[get_column_letter(i) + str(j)]=0
        ws_new[get_column_letter(i)+str(1)]=predmet_v_stolpcu
        ws_new["A"+str(j)]=sr["A"+str(j)].value
        ws_new["B" + str(j)]=sr["B"+str(j)].value
        ws_new["C" + str(j)]=sr["C"+str(j)].value


wb_new_two.save(dest)