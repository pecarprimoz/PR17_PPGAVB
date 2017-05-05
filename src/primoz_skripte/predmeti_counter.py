from collections import defaultdict
from openpyxl import load_workbook
predmeti = defaultdict(int)
from openpyxl import Workbook
wb = load_workbook("../podatki/izbirci_correct.xlsx")
sr= wb["correcttxt_new"]
c=2
predmeti_na_solo=defaultdict(set)
while c!=77747:
    sola =sr['B'+str(c)].value
    predmet=sr['C'+str(c)].value
    predmeti_na_solo[sola].add(predmet)
    c+=1
print(predmeti_na_solo)
predmeti_pravilno=defaultdict(int)
for key,value in predmeti_na_solo.items():
    for predmet in list(value):
        predmeti_pravilno[predmet] += 1
print(predmeti_pravilno)
wb_new = Workbook()

dest="../nasi_podatki/primoz/statistika_predmetov.xlsx"
ws1= wb_new.active
ws1.title = "St. predmetov nasploh"
ws1['A1']="Predmet"
ws1['B1']="Pojavitev"
counter=2
for key,value in predmeti_pravilno.items():
    ws1['A'+str(counter)]=key
    ws1['B'+str(counter)]=value
    print(key,value)
    counter+=1
wb_new.save(filename=dest)
