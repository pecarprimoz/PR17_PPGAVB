from collections import defaultdict
from openpyxl import load_workbook
predmeti = defaultdict(int)
from openpyxl import Workbook
wb = load_workbook("../../nasi_podatki/primoz/oddelki_otroc.xlsx")
sr= wb["List1"]
leto_ucenci=defaultdict(int)
leto_ucenke=defaultdict(int)
leto_romi=defaultdict(int)

#109541
#12 - '1. razred'
#13 - '2. razred'
#14 - '3. razred'
#15 - '4. razred'
#16 - '5. razred'
#17 - '6. razred'
#18 - '7. razred'
#19 - '8. razred'
#20 - '9. razred'

print(leto_ucenci)
print(leto_ucenke)
print(leto_romi)

wb_new = Workbook()

dest="../../nasi_podatki/primoz/oddelki_otroc_pravilni.xlsx"
ws1= wb_new.active
ws1.title = "St. otroc po odelkih"
ws1['A1']="Leto"
ws1['B1']="Oddelek"
ws1['C1']="Ucenci"
ws1['D1']="Ucenke"
ws1['E1']="Romi"
counter=2
while counter!=85764:
    razred = sr["A" + str(counter)].value
    oddelek = sr["B" + str(counter)].value
    ucenci = sr["C" + str(counter)].value
    ucenke = sr["D" + str(counter)].value
    romi = sr["E" + str(counter)].value
    if razred==12:
        ws1['A' + str(counter)]="1. razred"
    elif razred==13:
        ws1['A' + str(counter)] = "2. razred"
    elif razred == 14:
        ws1['A' + str(counter)] = "3. razred"
    elif razred == 15:
        ws1['A' + str(counter)] = "4. razred"
    elif razred == 16:
        ws1['A' + str(counter)] = "5. razred"
    elif razred == 17:
        ws1['A' + str(counter)] = "6. razred"
    elif razred == 18:
        ws1['A' + str(counter)] = "7. razred"
    elif razred == 19:
        ws1['A' + str(counter)] = "8. razred"
    elif razred == 20:
        ws1['A' + str(counter)] = "9. razred"
    ws1["B"+str(counter)]=oddelek
    ws1["C" + str(counter)] = ucenci
    ws1["D" + str(counter)] = ucenke
    ws1["E" + str(counter)] = romi
    counter+=1

wb_new.save(filename=dest)